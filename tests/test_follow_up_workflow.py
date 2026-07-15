import os
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

import main


class DummyRequest:
    def url_for(self, name, **kwargs):
        return f"http://testserver/{name}"


class FollowUpWorkflowTests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.xlsx_path = os.path.join(self.temp_dir.name, "complaints.xlsx")
        self.original_path = main.COMPLAINTS_FILE
        main.COMPLAINTS_FILE = self.xlsx_path
        main.init_complaints_workbook()

        wb = main.load_complaints_workbook()
        ws = wb["Complaints"]
        ws.append([
            "CMP-0001",
            "2026-01-01 10:00:00",
            "Test User",
            "user@mehtagroup.com",
            "Laptop",
            "IT",
            "L1",
            "Original issue",
            "Initial description",
            "RE1",
            "2026-01-01 10:05:00",
            "Original resolution",
            "2026-01-01 10:10:00",
            "Follow-up issue 1",
            "2026-01-01 10:15:00",
            "Solved follow-up 1",
            "2026-01-01 10:16:00",
            "Follow-up issue 2",
            "2026-01-01 10:17:00",
            "",
            "",
            "Reopened",
            ""  # Closing Time
        ])
        main.save_complaints_workbook(wb)
        wb.close()

    def tearDown(self):
        main.COMPLAINTS_FILE = self.original_path
        self.temp_dir.cleanup()

    def test_engineer_extra_solution_does_not_overwrite_main_resolution(self):
        token = main.generate_token({
            "complaint_id": "CMP-0001",
            "role": "engineer_action",
            "engineer_code": "RE1",
        })

        with patch.object(main, "send_email", return_value=True):
            response = main.engineer_reopen_submit(
                request=DummyRequest(),
                token=token,
                solution="Solved follow-up 2",
            )

        self.assertEqual(response.status_code, 200)

        wb = main.load_complaints_workbook()
        ws = wb["Complaints"]
        row = ws.max_row
        status_value = ws.cell(row=row, column=22).value
        resolution_value = ws.cell(row=row, column=12).value
        extra_solution_2_value = ws.cell(row=row, column=20).value

        self.assertEqual(status_value, "Waiting for Customer")
        self.assertEqual(resolution_value, "Original resolution")
        self.assertEqual(extra_solution_2_value, "Solved follow-up 2")

    def test_customer_follow_up_after_limit_closes_complaint_without_engineer_link(self):
        wb = main.load_complaints_workbook()
        ws = wb["Complaints"]
        ws.cell(row=5, column=16).value = "Solved follow-up 1"
        ws.cell(row=5, column=17).value = "2026-01-01 10:16:00"
        ws.cell(row=5, column=20).value = "Solved follow-up 2"
        ws.cell(row=5, column=21).value = "2026-01-01 10:17:00"
        ws.cell(row=5, column=22).value = "Reopened"
        main.save_complaints_workbook(wb)
        wb.close()

        complaint = main.get_complaint_by_id("CMP-0001")
        token = main.generate_token({
            "complaint_id": complaint["complaint_id"],
            "role": "customer_ack",
        })

        with patch.object(main, "send_email", return_value=True) as mocked_send_email:
            response = main.acknowledge_no(
                request=DummyRequest(),
                complaint_id="CMP-0001",
                token=token,
                follow_up_issue="Still not resolved",
            )

        self.assertEqual(response.status_code, 200)

        updated = main.get_complaint_by_id("CMP-0001")
        self.assertEqual(updated["status"], "Closed")

        sent_mail_bodies = [call.args[2] for call in mocked_send_email.call_args_list]
        self.assertTrue(sent_mail_bodies)
        self.assertNotIn("http", "\n".join(sent_mail_bodies))
        self.assertIn("contact support directly", "\n".join(sent_mail_bodies).lower())

    def test_auto_close_pending_complaints(self):
        from datetime import datetime, timedelta
        import sys
        from io import StringIO

        wb = main.load_complaints_workbook()
        ws = wb["Complaints"]
        # Clear existing data rows except header
        while ws.max_row > 4:
            ws.delete_rows(ws.max_row)

        four_days_ago = (datetime.now() - timedelta(days=4)).strftime("%Y-%m-%d %H:%M:%S")
        two_days_ago = (datetime.now() - timedelta(days=2)).strftime("%Y-%m-%d %H:%M:%S")

        ws.append([
            "CMP-0002", "2026-01-01 10:00:00", "User A", "usera@mehtagroup.com",
            "Laptop", "IT", "L1", "Issue", "Desc", "RE1", "2026-01-01 10:05:00",
            "Resolution", four_days_ago, "", "", "", "", "", "", "", "",
            "Waiting for Customer", ""
        ])

        ws.append([
            "CMP-0003", "2026-01-01 10:00:00", "User B", "userb@mehtagroup.com",
            "Laptop", "IT", "L2", "Issue", "Desc", "RE2", "2026-01-01 10:05:00",
            "Resolution", two_days_ago, "", "", "", "", "", "", "", "",
            "Waiting for Customer", ""
        ])

        table = ws.tables["ComplaintsTable"]
        table.ref = f"A4:W{ws.max_row}"

        main.save_complaints_workbook(wb)
        wb.close()

        captured_output = StringIO()
        sys.stdout = captured_output
        try:
            with patch.object(main, "send_email", return_value=True) as mocked_send_email:
                main.auto_close_pending_complaints()
        finally:
            sys.stdout = sys.__stdout__

        comp2 = main.get_complaint_by_id("CMP-0002")
        comp3 = main.get_complaint_by_id("CMP-0003")

        self.assertEqual(comp2["status"], "Closed")
        self.assertNotEqual(comp2["closing_time"], "")
        self.assertEqual(comp3["status"], "Waiting for Customer")
        self.assertEqual(comp3["closing_time"], "")

        output_str = captured_output.getvalue()
        self.assertIn("CMP-0002 status changed to Closed (Auto Closed)", output_str)
        self.assertNotIn("CMP-0003", output_str)

    def test_assign_engineer_post_redirects_with_token(self):
        token = main.generate_token({
            "complaint_id": "CMP-0001",
            "role": "admin_assign",
        })

        with patch.object(main, "send_email", return_value=True):
            response = main.assign_engineer_post(
                request=DummyRequest(),
                token=token,
                engineer_code="RE1",
            )

        self.assertEqual(response.status_code, 303)
        self.assertIn("token=", response.headers["location"])

    @patch('main.load_device_data')
    def test_submit_form_enforces_domain(self, mock_load):
        mock_load.return_value = [
            {"device_type": "Laptop", "department": "IT", "device_id": "L1"}
        ]

        with patch.object(main, "send_email", return_value=True):
            # Test non-mehtagroup.com domain
            response = main.submit_form(
                request=DummyRequest(),
                name="Test User",
                email="user@gmail.com",
                device_type="Laptop",
                department="IT",
                device_id="L1",
                issue="Keyboard",
                custom_issue="",
                description="Test description",
                edit_id=""
            )
            self.assertEqual(getattr(response, "status_code", 200), 200)

            # Test mehtagroup.com domain
            response_ok = main.submit_form(
                request=DummyRequest(),
                name="Test User",
                email="user@mehtagroup.com",
                device_type="Laptop",
                department="IT",
                device_id="L1",
                issue="Keyboard",
                custom_issue="",
                description="Test description",
                edit_id=""
            )
            self.assertEqual(response_ok.status_code, 303)


if __name__ == "__main__":
    unittest.main()
