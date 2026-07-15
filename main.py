from fastapi import FastAPI, Request, Form
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse
from dotenv import load_dotenv
load_dotenv()
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
import pandas as pd
import json
import os
import smtplib
import io
import msoffcrypto
import ssl
from email.mime.text import MIMEText
from datetime import datetime, timedelta
from contextlib import asynccontextmanager
from urllib.parse import quote_plus
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.interval import IntervalTrigger

scheduler = BackgroundScheduler()

@asynccontextmanager
async def lifespan(app: FastAPI):
    scheduler.add_job(
        auto_close_pending_complaints,
        trigger=IntervalTrigger(minutes=30),
        id="auto_close_pending_complaints",
        replace_existing=True,
    )
    scheduler.start()
    try:
        yield
    finally:
        scheduler.shutdown()
    

app = FastAPI(lifespan=lifespan)
templates = Jinja2Templates(directory="templates")


EXCEL_FILE = os.path.abspath("RNV-DESKTOP-DETAILS.xlsx")
COMPLAINTS_FILE = os.path.abspath("complaints.xlsx")


ADMIN_EMAILS = [
    email.strip()
    for email in os.getenv("ADMIN_EMAILS", "rcmotivaras@mehtagroup.com,ybgupta@mehtagroup.com").split(",")
    if email.strip()
]
EXCEL_ADMIN_PASSWORD = os.getenv("EXCEL_ADMIN_PASSWORD", "RNVAdmin@2026")


RE_MAP = {
    "RE1": "Resolving Engineer 1",
    "RE2": "Resolving Engineer 2",
    "RE3": "Resolving Engineer 3",
}

ENGINEER_EMAIL_MAP = {
    "RE1": os.getenv("RE1_EMAIL", "hmparmar@mehtagroup.com"),
    "RE2": os.getenv("RE2_EMAIL", "abbhadaraka@mehtagroup.com"),
    "RE3": os.getenv("RE3_EMAIL", "vbmori@mehtagroup.com"),
}


SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_EMAIL = os.getenv("SMTP_EMAIL", "uchitur3@gmail.com")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "onki wjnb lswd ebsb")


APP_SECRET_KEY = os.getenv("APP_SECRET_KEY", "change-this-secret-key")
TOKEN_SALT = "complaint-routing-flow"

DEFAULT_ISSUES = [
    "Mouse",
    "Keyboard",
    "CPU",
    "Laptop",
    "Monitor",
    "Printer",
    "Wifi ",
    "Drive related issues",
    "SAP related issues",
    "Webcam/Headset",
    "Email related issues",
    "CCTV related issues",
    "Other issues"
]
CUSTOM_ISSUES = set()


serializer = URLSafeTimedSerializer(APP_SECRET_KEY)


def generate_token(data: dict):
    return serializer.dumps(data, salt=TOKEN_SALT)


def verify_token(token: str, max_age: int = 24 * 3600):
    return serializer.loads(token, salt=TOKEN_SALT, max_age=max_age)


def send_email(to_email, subject, body):
    try:
        msg = MIMEText(body, "plain", "utf-8")
        msg["Subject"] = subject
        msg["From"] = SMTP_EMAIL
        msg["To"] = to_email

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=30) as server:
            server.ehlo()
            server.starttls(context=ssl.create_default_context())
            server.ehlo()
            server.login(SMTP_EMAIL, SMTP_PASSWORD)
            server.send_message(msg)

        print(f"Mail sent successfully to {to_email}")
        return True

    except Exception as e:
        print("Email sending failed:", str(e))
        return False

RNV_DESKTOP_EXCEL_PASSWORD = "RNVAdmin"

def read_password_protected_excel(file_path: str, password: str, sheet_name="HARDWARE", header=None):
    decrypted = io.BytesIO()

    with open(file_path, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)

    decrypted.seek(0)
    return pd.read_excel(decrypted, sheet_name=sheet_name, header=header)


def load_complaints_workbook() -> Workbook:
    if not os.path.exists(COMPLAINTS_FILE):
        init_complaints_workbook()

    with open(COMPLAINTS_FILE, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        is_encrypted = office_file.is_encrypted()

    if is_encrypted:
        decrypted = io.BytesIO()
        with open(COMPLAINTS_FILE, "rb") as f:
            office_file = msoffcrypto.OfficeFile(f)
            office_file.load_key(password=EXCEL_ADMIN_PASSWORD)
            office_file.decrypt(decrypted)
        decrypted.seek(0)
        return load_workbook(decrypted)
    else:
        return load_workbook(COMPLAINTS_FILE)


def save_complaints_workbook(wb: Workbook):
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from msoffcrypto.format.ooxml import OOXMLFile
    file = OOXMLFile(buffer)

    temp_file = COMPLAINTS_FILE + ".tmp"
    try:
        with open(temp_file, "wb") as out:
            file.encrypt(EXCEL_ADMIN_PASSWORD, out)
        os.replace(temp_file, COMPLAINTS_FILE)
    except Exception as e:
        if os.path.exists(temp_file):
            os.remove(temp_file)
        raise e


def read_complaints_df(sheet_name="Complaints", header=3) -> pd.DataFrame:
    with open(COMPLAINTS_FILE, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        is_encrypted = office_file.is_encrypted()

    if is_encrypted:
        decrypted = io.BytesIO()
        with open(COMPLAINTS_FILE, "rb") as f:
            office_file = msoffcrypto.OfficeFile(f)
            office_file.load_key(password=EXCEL_ADMIN_PASSWORD)
            office_file.decrypt(decrypted)
        decrypted.seek(0)
        return pd.read_excel(decrypted, sheet_name=sheet_name, header=header)
    else:
        return pd.read_excel(COMPLAINTS_FILE, sheet_name=sheet_name, header=header)


def load_device_data():
    df = read_password_protected_excel(
        EXCEL_FILE,
        password=RNV_DESKTOP_EXCEL_PASSWORD,
        sheet_name="HARDWARE",
        header=None
    )

    rows = df.fillna("").values.tolist()

    device_data = []
    current_type = None
    header_map = {}

    def is_section_header(values, keyword):
        nonempty = [v for v in values if v]
        return len(nonempty) == 1 and nonempty[0].upper() == keyword

    for row in rows:
        values = [str(x).strip() for x in row]

        if is_section_header(values, "DESKTOP"):
            current_type = "Desktop"
            header_map = {}
            continue
        if is_section_header(values, "LAPTOPS"):
            current_type = "Laptop"
            header_map = {}
            continue
        if is_section_header(values, "SERVER"):
            current_type = "Server"
            header_map = {}
            continue

        if "DEPPT." in values and "COMPUTER NO." in values and "STATUS" in values:
            header_map = {
                "department": values.index("DEPPT."),
                "device_id": values.index("COMPUTER NO."),
                "type": values.index("TYPE") if "TYPE" in values else None,
            }
            continue

        if current_type and header_map:
            department = str(row[header_map["department"]]).strip()
            device_id = str(row[header_map["device_id"]]).strip()

            detected_type = current_type
            if header_map.get("type") is not None:
                raw_type = str(row[header_map["type"]]).strip().lower()
                if raw_type == "laptop":
                    detected_type = "Laptop"
                elif raw_type == "desktop":
                    detected_type = "Desktop"
                elif "server" in raw_type:
                    detected_type = "Server"

            if (
                department
                and device_id
                and department != "DEPPT."
                and device_id != "COMPUTER NO."
                and "RNVSPARE" not in device_id.upper()
            ):
                device_data.append({
                    "device_type": detected_type,
                    "department": department,
                    "device_id": device_id
                })

    return device_data


def get_departments(data, device_type):
    return sorted({
        row["department"].strip()
        for row in data
        if row["device_type"] == device_type and row["department"].strip()
    })


def get_devices(data, device_type, department):
    return sorted(
        [
            row for row in data
            if row["device_type"] == device_type and row["department"].strip() == department.strip()
        ],
        key=lambda x: x["device_id"]
    )


def is_valid_admin_password(password: str) -> bool:
    return str(password or "").strip() == EXCEL_ADMIN_PASSWORD


def get_issue_options():
    base_issues = [item for item in DEFAULT_ISSUES if item != "Other issues"]
    custom_issue_values = [item for item in sorted(CUSTOM_ISSUES, key=lambda item: item.lower()) if item != "Other issues"]
    all_issues = base_issues + custom_issue_values

    if "Other issues" in DEFAULT_ISSUES:
        all_issues.append("Other issues")

    return all_issues


def init_complaints_workbook():
    if os.path.exists(COMPLAINTS_FILE):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Complaints"

    ws["A1"] = "Complaint Register"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14)
    ws["A2"] = "Generated from FastAPI complaint form"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10)

    headers = [
        "Complaint ID",
        "Timestamp",
        "Name",
        "Email",
        "Device Type",
        "Department",
        "Device ID",
        "Issue",
        "Description",
        "Assigned Engineer",
        "Assigned Timestamp",
        "Resolution",
        "Resolution Timestamp",
        "Extra Issue 1",
        "Extra Issue 1 Timestamp",
        "Extra Solution 1",
        "Extra Solution 1 Timestamp",
        "Extra Issue 2",
        "Extra Issue 2 Timestamp",
        "Extra Solution 2",
        "Extra Solution 2 Timestamp",
        "Status",
        "Closing Time"
    ]

    header_row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [15, 20, 22, 28, 15, 20, 18, 25, 45, 20, 22, 30, 22, 28, 22, 30, 22, 28, 22, 30, 22, 15, 22]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    table = Table(displayName="ComplaintsTable", ref="A4:W5")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(table)
    ws.freeze_panes = "A5"

    save_complaints_workbook(wb)
    wb.close()


def update_last_modified(ws, row_num):
    pass

def get_next_complaint_id(ws):
    if ws.max_row <= 4:
        return "CMP-0001"

    last_id = ws.cell(ws.max_row, 1).value
    if not last_id or not str(last_id).startswith("CMP-"):
        return "CMP-0001"

    last_num = int(str(last_id).split("-")[1])
    return f"CMP-{last_num + 1:04d}"


def save_complaint(name, email, device_type, department, device_id, issue, description):
    init_complaints_workbook()
    wb = None

    try:
        wb = load_complaints_workbook()
        ws = wb["Complaints"]

        complaint_id = get_next_complaint_id(ws)
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        ws.append([
            complaint_id,
            now_str,
            name,
            email,
            device_type,
            department,
            device_id,
            issue,
            description,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "Open",
            ""  # Closing Time
        ])

        table = ws.tables["ComplaintsTable"]
        table.ref = f"A4:W{ws.max_row}"
        if getattr(table, "autoFilter", None):
            table.autoFilter.ref = table.ref

        save_complaints_workbook(wb)
        wb.close()
        return complaint_id

    except PermissionError:
        if wb:
            try:
                wb.close()
            except Exception:
                pass
        raise Exception("complaints.xlsx is open or locked. Please close it and try again.")


def update_complaint_in_excel(complaint_id, name, email, device_type, department, device_id, issue, description):
    if not os.path.exists(COMPLAINTS_FILE):
        return False

    wb = None
    try:
        wb = load_complaints_workbook()
        ws = wb["Complaints"]

        headers = {}
        for col in range(1, ws.max_column + 1):
            headers[ws.cell(row=4, column=col).value] = col

        id_col = headers.get("Complaint ID")
        name_col = headers.get("Name")
        email_col = headers.get("Email")
        device_type_col = headers.get("Device Type")
        department_col = headers.get("Department")
        device_id_col = headers.get("Device ID")
        issue_col = headers.get("Issue")
        description_col = headers.get("Description")

        updated = False
        for r in range(5, ws.max_row + 1):
            cell_val = str(ws.cell(row=r, column=id_col).value).strip() if id_col else ""
            if cell_val == complaint_id.strip():
                if name_col: ws.cell(row=r, column=name_col).value = name
                if email_col: ws.cell(row=r, column=email_col).value = email
                if device_type_col: ws.cell(row=r, column=device_type_col).value = device_type
                if department_col: ws.cell(row=r, column=department_col).value = department
                if device_id_col: ws.cell(row=r, column=device_id_col).value = device_id
                if issue_col: ws.cell(row=r, column=issue_col).value = issue
                if description_col: ws.cell(row=r, column=description_col).value = description
                updated = True
                break

        if updated:
            save_complaints_workbook(wb)

        wb.close()
        return updated

    except PermissionError:
        if wb:
            try:
                wb.close()
            except Exception:
                pass
        raise Exception("complaints.xlsx is open or locked. Please close it and try again.")


def get_complaint_by_id(complaint_id):
    if not os.path.exists(COMPLAINTS_FILE):
        return None

    df = read_complaints_df(sheet_name="Complaints", header=3)
    df = df.fillna("")
    cols = list(df.columns)

    for _, row in df.iterrows():
        if str(row.get("Complaint ID", "")).strip() == complaint_id.strip():
            return {
                "complaint_id": str(row.get("Complaint ID", "")).strip(),
                "timestamp": str(row.get("Timestamp", "")).strip(),
                "name": str(row.get("Name", "")).strip(),
                "email": str(row.get("Email", "")).strip(),
                "device_type": str(row.get("Device Type", "")).strip(),
                "department": str(row.get("Department", "")).strip(),
                "device_id": str(row.get("Device ID", "")).strip(),
                "issue": str(row.get("Issue", "")).strip(),
                "description": str(row.get("Description", "")).strip(),
                "assigned_engineer": str(row.get("Assigned Engineer", "")).strip() if "Assigned Engineer" in cols else "",
                "assigned_timestamp": str(row.get("Assigned Timestamp", "")).strip() if "Assigned Timestamp" in cols else "",
                "resolution": str(row.get("Resolution", "")).strip() if "Resolution" in cols else "",
                "resolution_timestamp": str(row.get("Resolution Timestamp", "")).strip() if "Resolution Timestamp" in cols else "",
                "extra_issue_1": str(row.get("Extra Issue 1", "")).strip() if "Extra Issue 1" in cols else "",
                "extra_issue_1_timestamp": str(row.get("Extra Issue 1 Timestamp", "")).strip() if "Extra Issue 1 Timestamp" in cols else "",
                "extra_solution_1": str(row.get("Extra Solution 1", "")).strip() if "Extra Solution 1" in cols else "",
                "extra_solution_1_timestamp": str(row.get("Extra Solution 1 Timestamp", "")).strip() if "Extra Solution 1 Timestamp" in cols else "",
                "extra_issue_2": str(row.get("Extra Issue 2", "")).strip() if "Extra Issue 2" in cols else "",
                "extra_issue_2_timestamp": str(row.get("Extra Issue 2 Timestamp", "")).strip() if "Extra Issue 2 Timestamp" in cols else "",
                "extra_solution_2": str(row.get("Extra Solution 2", "")).strip() if "Extra Solution 2" in cols else "",
                "extra_solution_2_timestamp": str(row.get("Extra Solution 2 Timestamp", "")).strip() if "Extra Solution 2 Timestamp" in cols else "",
                "status": str(row.get("Status", "")).strip(),
                "closing_time": str(row.get("Closing Time", "")).strip() if "Closing Time" in cols else "",
            }

    return None

def save_extra_issue(complaint_id, follow_up_issue):
    if not os.path.exists(COMPLAINTS_FILE):
        return False, "Complaint file not found."

    wb = None
    try:
        wb = load_complaints_workbook()
        ws = wb["Complaints"]

        headers = {}
        for col in range(1, ws.max_column + 1):
            headers[ws.cell(row=4, column=col).value] = col

        id_col = headers.get("Complaint ID")
        status_col = headers.get("Status")
        extra1_col = headers.get("Extra Issue 1")
        extra1_time_col = headers.get("Extra Issue 1 Timestamp")
        extra2_col = headers.get("Extra Issue 2")
        extra2_time_col = headers.get("Extra Issue 2 Timestamp")

        for r in range(5, ws.max_row + 1):
            cell_val = str(ws.cell(row=r, column=id_col).value or "").strip()
            if cell_val == complaint_id.strip():
                extra1_val = str(ws.cell(row=r, column=extra1_col).value or "").strip() if extra1_col else ""
                extra2_val = str(ws.cell(row=r, column=extra2_col).value or "").strip() if extra2_col else ""
                now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                if not extra1_val:
                    ws.cell(row=r, column=extra1_col).value = follow_up_issue.strip()
                    if extra1_time_col:
                        ws.cell(row=r, column=extra1_time_col).value = now_str
                elif not extra2_val:
                    ws.cell(row=r, column=extra2_col).value = follow_up_issue.strip()
                    if extra2_time_col:
                        ws.cell(row=r, column=extra2_time_col).value = now_str
                else:
                    wb.close()
                    return False, "Maximum extra issues reached."

                if status_col:
                    ws.cell(row=r, column=status_col).value = "Reopened"

                save_complaints_workbook(wb)
                wb.close()
                return True, "Extra issue saved successfully."

        wb.close()
        return False, "Complaint ID not found."

    except PermissionError:
        if wb:
            try:
                wb.close()
            except Exception:
                pass
        raise Exception("complaints.xlsx is open or locked. Please close it and try again.")


def save_extra_solution(complaint_id, solution_text, status="Closed"):
    if not os.path.exists(COMPLAINTS_FILE):
        return False, "Complaint file not found."

    wb = None
    try:
        wb = load_complaints_workbook()
        ws = wb["Complaints"]

        headers = {}
        for col in range(1, ws.max_column + 1):
            headers[ws.cell(row=4, column=col).value] = col

        id_col = headers.get("Complaint ID")
        status_col = headers.get("Status")
        closing_time_col = headers.get("Closing Time")

        extra1_issue_col = headers.get("Extra Issue 1")
        extra1_solution_col = headers.get("Extra Solution 1")
        extra1_solution_time_col = headers.get("Extra Solution 1 Timestamp")

        extra2_issue_col = headers.get("Extra Issue 2")
        extra2_solution_col = headers.get("Extra Solution 2")
        extra2_solution_time_col = headers.get("Extra Solution 2 Timestamp")

        for r in range(5, ws.max_row + 1):
            cell_val = str(ws.cell(row=r, column=id_col).value or "").strip()
            if cell_val == complaint_id.strip():
                extra1_issue = str(ws.cell(row=r, column=extra1_issue_col).value or "").strip() if extra1_issue_col else ""
                extra1_solution = str(ws.cell(row=r, column=extra1_solution_col).value or "").strip() if extra1_solution_col else ""
                extra2_issue = str(ws.cell(row=r, column=extra2_issue_col).value or "").strip() if extra2_issue_col else ""
                extra2_solution = str(ws.cell(row=r, column=extra2_solution_col).value or "").strip() if extra2_solution_col else ""

                now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                if extra1_issue and not extra1_solution:
                    ws.cell(row=r, column=extra1_solution_col).value = solution_text.strip()
                    if extra1_solution_time_col:
                        ws.cell(row=r, column=extra1_solution_time_col).value = now_str
                elif extra2_issue and not extra2_solution:
                    ws.cell(row=r, column=extra2_solution_col).value = solution_text.strip()
                    if extra2_solution_time_col:
                        ws.cell(row=r, column=extra2_solution_time_col).value = now_str
                else:
                    wb.close()
                    return False, "No pending extra issue available for extra solution."

                if status_col:
                    ws.cell(row=r, column=status_col).value = status
                    if status == "Closed":
                        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Complaint {complaint_id} status changed to Closed")
                        if closing_time_col:
                            ws.cell(row=r, column=closing_time_col).value = now_str
                    else:
                        if closing_time_col:
                            ws.cell(row=r, column=closing_time_col).value = ""

                save_complaints_workbook(wb)
                wb.close()
                return True, "Extra solution saved successfully."

        wb.close()
        return False, "Complaint ID not found."

    except PermissionError:
        if wb:
            try:
                wb.close()
            except Exception:
                pass
        raise Exception("complaints.xlsx is open or locked. Please close it and try again.")

def update_assignment(complaint_id, engineer_code):
    if not os.path.exists(COMPLAINTS_FILE):
        return False

    wb = None
    try:
        wb = load_complaints_workbook()
        ws = wb["Complaints"]

        headers = {}
        for col in range(1, ws.max_column + 1):
            headers[ws.cell(row=4, column=col).value] = col

        id_col = headers.get("Complaint ID")
        assigned_col = headers.get("Assigned Engineer")
        assigned_time_col = headers.get("Assigned Timestamp")
        status_col = headers.get("Status")

        updated = False
        for r in range(5, ws.max_row + 1):
            cell_val = str(ws.cell(row=r, column=id_col).value).strip() if id_col else ""
            if cell_val == complaint_id.strip():
                if assigned_col:
                    ws.cell(row=r, column=assigned_col).value = engineer_code
                if assigned_time_col:
                    ws.cell(row=r, column=assigned_time_col).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                if status_col:
                    ws.cell(row=r, column=status_col).value = "Assigned"
                updated = True
                break

        if updated:
            save_complaints_workbook(wb)

        wb.close()
        return updated

    except PermissionError:
        if wb:
            try:
                wb.close()
            except Exception:
                pass
        raise Exception("complaints.xlsx is open or locked. Please close it and try again.")


def resolve_in_excel(complaint_id, status, solution):
    if not os.path.exists(COMPLAINTS_FILE):
        return False

    wb = None
    try:
        wb = load_complaints_workbook()
        ws = wb["Complaints"]

        headers = {}
        for col in range(1, ws.max_column + 1):
            headers[ws.cell(row=4, column=col).value] = col

        id_col = headers.get("Complaint ID")
        status_col = headers.get("Status")
        resolution_col = headers.get("Resolution")
        resolution_time_col = headers.get("Resolution Timestamp")
        closing_time_col = headers.get("Closing Time")

        updated = False
        for r in range(5, ws.max_row + 1):
            cell_val = str(ws.cell(row=r, column=id_col).value).strip() if id_col else ""
            if cell_val == complaint_id.strip():
                if status_col:
                    ws.cell(row=r, column=status_col).value = status
                    if status == "Closed":
                        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Complaint {complaint_id} status changed to Closed")
                        if closing_time_col:
                            ws.cell(row=r, column=closing_time_col).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        if closing_time_col:
                            ws.cell(row=r, column=closing_time_col).value = ""
                if resolution_col:
                    ws.cell(row=r, column=resolution_col).value = solution
                if resolution_time_col:
                    ws.cell(row=r, column=resolution_time_col).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                updated = True
                break

        if updated:
            save_complaints_workbook(wb)

        wb.close()
        return updated

    except PermissionError:
        if wb:
            try:
                wb.close()
            except Exception:
                pass
        raise Exception("complaints.xlsx is open or locked. Please close it and try again.")
def mark_waiting_for_customer(complaint_id, solution):
    if not os.path.exists(COMPLAINTS_FILE):
        return False

    wb = None
    try:
        wb = load_complaints_workbook()
        ws = wb["Complaints"]

        headers = {}
        for col in range(1, ws.max_column + 1):
            headers[ws.cell(row=4, column=col).value] = col

        id_col = headers.get("Complaint ID")
        status_col = headers.get("Status")
        resolution_col = headers.get("Resolution")
        resolution_time_col = headers.get("Resolution Timestamp")
        waiting_since_col = headers.get("Waiting Since")
        feedback_col = headers.get("Customer Feedback")
        followup_col = headers.get("Follow-up Issue")

        updated = False
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for r in range(5, ws.max_row + 1):
            cell_val = str(ws.cell(row=r, column=id_col).value).strip() if id_col else ""
            if cell_val == complaint_id.strip():
                if status_col:
                    ws.cell(row=r, column=status_col).value = "Waiting for Customer"
                if resolution_col:
                    ws.cell(row=r, column=resolution_col).value = solution
                if resolution_time_col:
                    ws.cell(row=r, column=resolution_time_col).value = now_str
                if waiting_since_col:
                    ws.cell(row=r, column=waiting_since_col).value = now_str
                if feedback_col:
                    ws.cell(row=r, column=feedback_col).value = ""
                if followup_col:
                    ws.cell(row=r, column=followup_col).value = ""
                updated = True
                break

        if updated:
            save_complaints_workbook(wb)

        wb.close()
        return updated

    except PermissionError:
        if wb:
            try:
                wb.close()
            except Exception:
                pass
        raise Exception("complaints.xlsx is open or locked. Please close it and try again.")

def update_customer_acknowledgement(complaint_id, feedback, follow_up_issue=""):
    if not os.path.exists(COMPLAINTS_FILE):
        return False

    wb = None
    try:
        wb = load_complaints_workbook()
        ws = wb["Complaints"]

        headers = {}
        for col in range(1, ws.max_column + 1):
            headers[ws.cell(row=4, column=col).value] = col

        id_col = headers.get("Complaint ID")
        status_col = headers.get("Status")
        feedback_col = headers.get("Customer Feedback")
        followup_col = headers.get("Follow-up Issue")
        waiting_since_col = headers.get("Waiting Since")
        closing_time_col = headers.get("Closing Time")

        updated = False
        for r in range(5, ws.max_row + 1):
            cell_val = str(ws.cell(row=r, column=id_col).value).strip() if id_col else ""
            if cell_val == complaint_id.strip():
                if feedback_col:
                    ws.cell(row=r, column=feedback_col).value = feedback

                if feedback == "Satisfied":
                    if status_col:
                        ws.cell(row=r, column=status_col).value = "Closed"
                        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Complaint {complaint_id} status changed to Closed")
                        if closing_time_col:
                            ws.cell(row=r, column=closing_time_col).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                elif feedback == "Not Satisfied":
                    if status_col:
                        ws.cell(row=r, column=status_col).value = "Reopened"
                    if followup_col:
                        ws.cell(row=r, column=followup_col).value = follow_up_issue.strip()
                    if closing_time_col:
                        ws.cell(row=r, column=closing_time_col).value = ""

                if waiting_since_col and feedback in ["Satisfied", "Not Satisfied"]:
                    ws.cell(row=r, column=waiting_since_col).value = ""

                updated = True
                break

        if updated:
            save_complaints_workbook(wb)

        wb.close()
        return updated

    except PermissionError:
        if wb:
            try:
                wb.close()
            except Exception:
                pass
        raise Exception("complaints.xlsx is open or locked. Please close it and try again.")

def auto_close_pending_complaints():
    if not os.path.exists(COMPLAINTS_FILE):
        return

    wb = None
    try:
        wb = load_complaints_workbook()
        ws = wb["Complaints"]

        headers = {}
        for col in range(1, ws.max_column + 1):
            headers[ws.cell(row=4, column=col).value] = col

        id_col = headers.get("Complaint ID")
        name_col = headers.get("Name")
        email_col = headers.get("Email")
        status_col = headers.get("Status")
        resolution_time_col = headers.get("Resolution Timestamp")
        extra1_sol_time_col = headers.get("Extra Solution 1 Timestamp")
        extra2_sol_time_col = headers.get("Extra Solution 2 Timestamp")
        waiting_since_col = headers.get("Waiting Since")
        closing_time_col = headers.get("Closing Time")

        changed = False
        now = datetime.now()

        for r in range(5, ws.max_row + 1):
            status = str(ws.cell(row=r, column=status_col).value).strip() if status_col else ""

            if status not in ["Waiting for Customer", "Resolved"]:
                continue

            # Gather all non-empty resolution/solution timestamps
            timestamps = []
            for col_idx in [resolution_time_col, extra1_sol_time_col, extra2_sol_time_col]:
                if col_idx:
                    val = str(ws.cell(row=r, column=col_idx).value or "").strip()
                    if val:
                        try:
                            dt = datetime.strptime(val, "%Y-%m-%d %H:%M:%S")
                            timestamps.append(dt)
                        except ValueError:
                            pass

            if not timestamps:
                continue

            latest_solve_dt = max(timestamps)

            if now - latest_solve_dt >= timedelta(days=3):
                complaint_id = str(ws.cell(row=r, column=id_col).value).strip() if id_col else ""
                user_name = str(ws.cell(row=r, column=name_col).value).strip() if name_col else ""
                user_email = str(ws.cell(row=r, column=email_col).value).strip() if email_col else ""

                ws.cell(row=r, column=status_col).value = "Closed"
                print(f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] Complaint {complaint_id} status changed to Closed (Auto Closed)")
                
                if closing_time_col:
                    ws.cell(row=r, column=closing_time_col).value = now.strftime("%Y-%m-%d %H:%M:%S")
                
                if waiting_since_col:
                    ws.cell(row=r, column=waiting_since_col).value = ""

                changed = True

                if user_email:
                    mail_subject = f"Complaint {complaint_id} auto-closed"
                    mail_body = f"""
Hello {user_name},

Your complaint has been automatically closed because we did not receive a response within 3 days.

Complaint ID: {complaint_id}
Status: Auto Closed

If the issue still persists, please raise the complaint again or contact support.

Regards,
Support Team
""".strip()
                    send_email(user_email, mail_subject, mail_body)

        if changed:
            save_complaints_workbook(wb)

        wb.close()

    except Exception as e:
        print("Auto-close job failed:", str(e))
        if wb:
            try:
                wb.close()
            except Exception:
                pass


@app.get("/", response_class=HTMLResponse)
def home(
    request: Request,
    name: str = "",
    email: str = "",
    device_type: str = "",
    department: str = "",
    device_id: str = "",
    issue: str = "",
    description: str = "",
    submitted: int = 0,
    complaint_id: str = "",
    edit_id: str = ""
):
    device_data = load_device_data()
    device_types = sorted({row["device_type"] for row in device_data})
    issues = get_issue_options()

    custom_issue = ""
    if edit_id:
        complaint = get_complaint_by_id(edit_id)
        if complaint:
            name = complaint.get("name", "")
            email = complaint.get("email", "")
            device_type = complaint.get("device_type", "")
            department = complaint.get("department", "")
            device_id = complaint.get("device_id", "")
            raw_issue = complaint.get("issue", "")
            description = complaint.get("description", "")
            
            if raw_issue in DEFAULT_ISSUES:
                issue = raw_issue
            else:
                issue = "Other issues"
                custom_issue = raw_issue

    departments = get_departments(device_data, device_type) if device_type else []
    devices = get_devices(device_data, device_type, department) if device_type and department else []

    return templates.TemplateResponse(
        request=request,
        name="form.html",
        context={
            "device_types": device_types,
            "departments": departments,
            "devices": devices,
            "selected_device_type": device_type,
            "selected_department": department,
            "selected_device_id": device_id,
            "name": name,
            "email": email,
            "issue": issue,
            "custom_issue": custom_issue,
            "description": description,
            "issues": issues,
            "submitted": submitted,
            "complaint_id": complaint_id,
            "edit_id": edit_id,
            "device_data_json": json.dumps(device_data),
            "error": ""
        }
    )


@app.post("/", response_class=HTMLResponse)
def submit_form(
    request: Request,
    name: str = Form(""),
    email: str = Form(""),
    device_type: str = Form(""),
    department: str = Form(""),
    device_id: str = Form(""),
    issue: str = Form(""),
    custom_issue: str = Form(""),
    description: str = Form(""),
    edit_id: str = Form("")
):
    complaint_id = ""
    device_data = load_device_data()
    device_types = sorted({row["device_type"] for row in device_data})
    issues = get_issue_options()

    departments = get_departments(device_data, device_type) if device_type else []
    devices = get_devices(device_data, device_type, department) if device_type and department else []

    selected_issue = issue.strip()
    if issue.strip() == "Other issues":
        selected_issue = custom_issue.strip() or "Other issues"

    if not all([
        name.strip(),
        email.strip(),
        device_type.strip(),
        department.strip(),
        device_id.strip(),
        selected_issue.strip(),
        description.strip()
    ]):
        return templates.TemplateResponse(
            request=request,
            name="form.html",
            context={
                "device_types": device_types,
                "departments": departments,
                "devices": devices,
                "selected_device_type": device_type,
                "selected_department": department,
                "selected_device_id": device_id,
                "name": name,
                "email": email,
                "issue": issue,
                "custom_issue": custom_issue,
                "description": description,
                "issues": issues,
                "submitted": 0,
                "complaint_id": "",
                "edit_id": edit_id,
                "device_data_json": json.dumps(device_data),
                "error": "Please fill all required fields."
            }
        )

    if not email.strip().lower().endswith("@mehtagroup.com"):
        return templates.TemplateResponse(
            request=request,
            name="form.html",
            context={
                "device_types": device_types,
                "departments": departments,
                "devices": devices,
                "selected_device_type": device_type,
                "selected_department": department,
                "selected_device_id": device_id,
                "name": name,
                "email": email,
                "issue": issue,
                "custom_issue": custom_issue,
                "description": description,
                "issues": issues,
                "submitted": 0,
                "complaint_id": "",
                "edit_id": edit_id,
                "device_data_json": json.dumps(device_data),
                "error": "Only @mehtagroup.com email addresses are allowed to register complaints."
            }
        )

    if selected_issue and selected_issue != "Other issues" and selected_issue not in DEFAULT_ISSUES:
        CUSTOM_ISSUES.add(selected_issue)

    try:
        if edit_id.strip():
            complaint_id = edit_id.strip()
            ok = update_complaint_in_excel(
                complaint_id=complaint_id,
                name=name,
                email=email,
                device_type=device_type,
                department=department,
                device_id=device_id,
                issue=selected_issue,
                description=description
            )
            if not ok:
                raise Exception("Complaint ID not found in Excel register.")
        else:
            complaint_id = save_complaint(
                name=name,
                email=email,
                device_type=device_type,
                department=department,
                device_id=device_id,
                issue=selected_issue,
                description=description
            )
    except Exception as e:
        return templates.TemplateResponse(
            request=request,
            name="form.html",
            context={
                "device_types": device_types,
                "departments": departments,
                "devices": devices,
                "selected_device_type": device_type,
                "selected_department": department,
                "selected_device_id": device_id,
                "name": name,
                "email": email,
                "issue": issue,
                "custom_issue": custom_issue,
                "description": description,
                "issues": issues,
                "submitted": 0,
                "complaint_id": "",
                "edit_id": edit_id,
                "device_data_json": json.dumps(device_data),
                "error": str(e)
            }
        )

    admin_token = generate_token({
        "complaint_id": complaint_id,
        "role": "admin_assign"
    })

    assign_link = str(request.url_for("assign_engineer")) + f"?token={admin_token}"

    if edit_id.strip():
        mail_subject = f"Complaint Edited - {complaint_id}"
        mail_body = f"""
Complaint {complaint_id} has been edited by the user.

Updated Details:
Name: {name}
Email: {email}
Device Type: {device_type}
Department: {department}
Device ID: {device_id}
Issue: {selected_issue}
Description: {description}

Open assignment page:
{assign_link}
""".strip()
    else:
        mail_subject = f"New Complaint Generated - {complaint_id}"
        mail_body = f"""
A new complaint has been generated.

Complaint ID: {complaint_id}
Name: {name}
Email: {email}
Device Type: {device_type}
Department: {department}
Device ID: {device_id}
Issue: {selected_issue}
Description: {description}

Open assignment page:
{assign_link}

Please assign the complaint to RE1, RE2 or RE3.
""".strip()

    for admin_email in ADMIN_EMAILS:
        send_email(admin_email, mail_subject, mail_body)

    return RedirectResponse(url=f"/submitted/{complaint_id}", status_code=303)


@app.get("/submitted/{complaint_id}", response_class=HTMLResponse)
def submission_success(request: Request, complaint_id: str):
    return templates.TemplateResponse(
        request=request,
        name="success.html",
        context={
            "complaint_id": complaint_id,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
    )
@app.get("/assigned/{complaint_id}", response_class=HTMLResponse)
def assignment_success(
    request: Request,
    complaint_id: str,
    engineer_code: str = "",
    engineer_email: str = "",
    token: str = ""
):
    complaint = get_complaint_by_id(complaint_id)
    assigned_time = ""
    if complaint:
        # try common keys that might store the assigned timestamp
        assigned_time = complaint.get("assigned_time") or complaint.get("assigned_timestamp") or complaint.get("Assigned Timestamp") or ""

    return templates.TemplateResponse(
        request=request,
        name="assign_success.html",
        context={
            "complaint_id": complaint_id,
            "engineer_code": engineer_code,
            "engineer_email": engineer_email,
            "assigned_at": assigned_time or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "token": token
        }
    )


@app.get("/resolved/{complaint_id}", response_class=HTMLResponse)
def resolution_success(
    request: Request,
    complaint_id: str,
    status: str = "",
    solution: str = "",
    token: str = ""
):
    return templates.TemplateResponse(
        request=request,
        name="resolve_success.html",
        context={
            "complaint_id": complaint_id,
            "status": status,
            "solution": solution,
            "resolved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "token": token
        }
    )

@app.get("/assign", response_class=HTMLResponse, name="assign_engineer")
def assign_engineer(request: Request, token: str = ""):
    try:
        payload = verify_token(token, max_age=24 * 3600)
    except SignatureExpired:
        return HTMLResponse("<h2>Assignment link expired</h2>", status_code=403)
    except BadSignature:
        return HTMLResponse("<h2>Invalid assignment token</h2>", status_code=403)

    if payload.get("role") != "admin_assign":
        return HTMLResponse("<h2>Unauthorized token role</h2>", status_code=403)

    complaint_id = payload.get("complaint_id", "")
    complaint = get_complaint_by_id(complaint_id)
    if not complaint:
        return HTMLResponse("<h2>Complaint not found</h2>", status_code=404)

    return templates.TemplateResponse(
        request=request,
        name="admin_access.html",
        context={
            "complaint_id": complaint_id,
            "complaint": complaint,
            "token": token,
            "error": "",
            "success": ""
        }
    )


@app.get("/assign-engineer", response_class=HTMLResponse, name="assign_engineer_page")
def assign_engineer_page(request: Request, token: str = ""):
    try:
        payload = verify_token(token, max_age=24 * 3600)
    except SignatureExpired:
        return HTMLResponse("<h2>Assignment link expired</h2>", status_code=403)
    except BadSignature:
        return HTMLResponse("<h2>Invalid assignment token</h2>", status_code=403)

    if payload.get("role") != "admin_assign":
        return HTMLResponse("<h2>Unauthorized token role</h2>", status_code=403)

    complaint_id = payload.get("complaint_id", "")
    complaint = get_complaint_by_id(complaint_id)
    if not complaint:
        return HTMLResponse("<h2>Complaint not found</h2>", status_code=404)

    return templates.TemplateResponse(
        request=request,
        name="assign_engineer.html",
        context={
            "complaint": complaint,
            "token": token,
            "error": "",
            "success": "",
            "selected_engineer": ""
        }
    )


@app.post("/admin_access", response_class=HTMLResponse, name="assign_access")
def assign_access_submit(
    request: Request,
    token: str = Form(...),
    password: str = Form("")
):
    try:
        payload = verify_token(token, max_age=24 * 3600)
    except SignatureExpired:
        return HTMLResponse("<h2>Assignment link expired</h2>", status_code=403)
    except BadSignature:
        return HTMLResponse("<h2>Invalid assignment token</h2>", status_code=403)

    if payload.get("role") != "admin_assign":
        return HTMLResponse("<h2>Unauthorized token role</h2>", status_code=403)

    complaint_id = payload.get("complaint_id", "")
    complaint = get_complaint_by_id(complaint_id)

    if not complaint:
        return HTMLResponse("<h2>Complaint not found</h2>", status_code=404)

    if not is_valid_admin_password(password):
        return templates.TemplateResponse(
            request=request,
            name="admin_access.html",
            context={
                "complaint_id": complaint_id,
                "complaint": complaint,
                "token": token,
                "error": "Incorrect password. Only admin can continue.",
                "success": ""
            }
        )

    return RedirectResponse(url=f"/assign-engineer?token={token}", status_code=303)


@app.post("/assign", response_class=HTMLResponse, name="assign_engineer_post")
def assign_engineer_post(
    request: Request,
    token: str = Form(...),
    engineer_code: str = Form(...)
):
    try:
        payload = verify_token(token, max_age=24 * 3600)
    except SignatureExpired:
        return HTMLResponse("<h2>Assignment link expired</h2>", status_code=403)
    except BadSignature:
        return HTMLResponse("<h2>Invalid assignment token</h2>", status_code=403)

    if payload.get("role") != "admin_assign":
        return HTMLResponse("<h2>Unauthorized token role</h2>", status_code=403)

    complaint_id = payload.get("complaint_id", "")
    complaint = get_complaint_by_id(complaint_id)

    if not complaint:
        return HTMLResponse("<h2>Complaint not found</h2>", status_code=404)

    if engineer_code not in RE_MAP:
        return templates.TemplateResponse(
            request=request,
            name="assign_engineer.html",
            context={
                "complaint": complaint,
                "token": token,
                "error": "Please select a valid engineer.",
                "success": "",
                "selected_engineer": engineer_code
            }
        )

    engineer_email = ENGINEER_EMAIL_MAP.get(engineer_code, "").strip()
    if not engineer_email:
        return templates.TemplateResponse(
            request=request,
            name="assign_engineer.html",
            context={
                "complaint": complaint,
                "token": token,
                "error": f"No email configured for {engineer_code}.",
                "success": "",
                "selected_engineer": engineer_code
            }
        )

    try:
        update_assignment(complaint_id, engineer_code)
    except Exception as e:
        return templates.TemplateResponse(
            request=request,
            name="assign_engineer.html",
            context={
                "complaint": complaint,
                "token": token,
                "error": str(e),
                "success": "",
                "selected_engineer": engineer_code
            }
        )

    engineer_token = generate_token({
        "complaint_id": complaint_id,
        "role": "engineer_action",
        "engineer_code": engineer_code
    })

    engineer_link = str(request.url_for("engineer_page")) + f"?token={engineer_token}"

    mail_subject = f"[{engineer_code}] Access for Complaint - {complaint_id}"
    mail_body = f"""
Complaint {complaint_id} has been assigned to {engineer_code}.

Open the complaint page below:
{engineer_link}
""".strip()

    send_email(engineer_email, mail_subject, mail_body)

    return RedirectResponse(
        url=f"/assigned/{complaint_id}?engineer_code={engineer_code}&engineer_email={engineer_email}&token={token}",
        status_code=303
    )


@app.get("/engineer", response_class=HTMLResponse, name="engineer_page")
def engineer_page(request: Request, token: str = ""):
    try:
        payload = verify_token(token, max_age=24 * 3600)
    except SignatureExpired:
        return HTMLResponse("<h2>Engineer link expired</h2>", status_code=403)
    except BadSignature:
        return HTMLResponse("<h2>Invalid engineer token</h2>", status_code=403)

    if payload.get("role") != "engineer_action":
        return HTMLResponse("<h2>Unauthorized token role</h2>", status_code=403)

    complaint_id = payload.get("complaint_id", "")
    engineer_code = payload.get("engineer_code", "")
    complaint = get_complaint_by_id(complaint_id)

    if not complaint:
        return HTMLResponse("<h2>Complaint not found</h2>", status_code=404)

    return templates.TemplateResponse(
        request=request,
        name="response.html",
        context={
            "complaint": complaint,
            "token": token,
            "engineer_code": engineer_code
        }
    )


@app.post("/engineer/reopen", response_class=HTMLResponse, name="engineer_reopen_submit")
def engineer_reopen_submit(request: Request, token: str = Form(...), solution: str = Form("")):
    try:
        payload = verify_token(token, max_age=24 * 3600)
    except SignatureExpired:
        return HTMLResponse("<h2>Engineer link expired</h2>", status_code=403)
    except BadSignature:
        return HTMLResponse("<h2>Invalid engineer token</h2>", status_code=403)

    if payload.get("role") != "engineer_action":
        return HTMLResponse("<h2>Unauthorized</h2>", status_code=403)

    complaint_id = payload.get("complaint_id", "")
    engineer_code = payload.get("engineer_code", "")

    complaint = get_complaint_by_id(complaint_id)
    if not complaint:
        return HTMLResponse("<h2>Complaint not found</h2>", status_code=404)

    # Save the engineer's solution against the current follow-up issue in Excel
    try:
        if complaint.get("extra_issue_1") and not complaint.get("extra_solution_1"):
            ok, msg = save_extra_solution(complaint_id, solution.strip(), status="Waiting for Customer")
            if not ok:
                return HTMLResponse(f"<h2>{msg}</h2>", status_code=500)
        elif complaint.get("extra_issue_2") and not complaint.get("extra_solution_2"):
            ok, msg = save_extra_solution(complaint_id, solution.strip(), status="Waiting for Customer")
            if not ok:
                return HTMLResponse(f"<h2>{msg}</h2>", status_code=500)
        else:
            # Fallback: preserve the main resolution and keep the complaint in waiting status
            mark_waiting_for_customer(complaint_id, complaint.get("resolution", ""))
    except Exception as e:
        return HTMLResponse(f"<h2>{str(e)}</h2>", status_code=500)

    # Reset customer acknowledgement and mark waiting for customer so the user can re-acknowledge
    try:
        mark_waiting_for_customer(complaint_id, complaint.get("resolution", ""))
    except Exception:
        pass

    # Notify customer of the engineer's update and include an acknowledgement link
    user_email = complaint.get("email", "")
    if user_email:
        customer_token = generate_token({
            "complaint_id": complaint_id,
            "role": "customer_ack"
        })
        acknowledgement_link = str(request.url_for("acknowledge_page", complaint_id=complaint_id)) + f"?token={customer_token}"

        mail_subject = f"Update on your complaint {complaint_id}"
        mail_body = f"""
Hello {complaint.get('name', '')},

Sorry for the trouble caused. Our engineer has provided an update on your complaint.

Complaint ID: {complaint_id}
Update: {solution.strip()}

Please confirm whether the issue is now resolved by clicking the link below:
{acknowledgement_link}

Regards,
Support Team
""".strip()
        send_email(user_email, mail_subject, mail_body)

    return templates.TemplateResponse(
        request=request,
        name="ack_result.html",
        context={
            "title": "Update sent",
            "message": "The update has been sent to the customer with an acknowledgement link."
        }
    )


@app.get("/engineer/reopen/{complaint_id}", response_class=HTMLResponse, name="engineer_reopen")
def engineer_reopen_page(request: Request, complaint_id: str, token: str = ""):
    complaint = get_complaint_by_id(complaint_id)
    if not complaint:
        return HTMLResponse("<h2>Complaint not found</h2>", status_code=404)

    if token:
        try:
            payload = verify_token(token, max_age=24 * 3600)
        except (SignatureExpired, BadSignature):
            return HTMLResponse("<h2>Engineer link expired or invalid</h2>", status_code=403)

        if payload.get("role") != "engineer_action" or payload.get("complaint_id") != complaint_id:
            return HTMLResponse("<h2>Unauthorized</h2>", status_code=403)

    return templates.TemplateResponse(
        request=request,
        name="reopen_engineer.html",
        context={
            "complaint": complaint,
            "token": token,
            "error": "",
            "success": ""
        }
    )


@app.get("/excel-access", response_class=HTMLResponse)
def excel_access_page(request: Request, error: str = ""):
    return templates.TemplateResponse(
        request=request,
        name="excel_access.html",
        context={"error": error}
    )


@app.post("/excel-access", response_class=HTMLResponse)
def excel_access_submit(request: Request, password: str = Form("")):
    if is_valid_admin_password(password):
        return RedirectResponse(url=f"/download-excel?password={password}", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="excel_access.html",
        context={"error": "Incorrect password. Only admin can use this page."}
    )


@app.get("/download-excel")
def download_excel(request: Request, password: str = ""):
    if not is_valid_admin_password(password):
        return RedirectResponse(url="/excel-access?error=unauthorized", status_code=303)

    if not os.path.exists(COMPLAINTS_FILE):
        init_complaints_workbook()

    return FileResponse(
        COMPLAINTS_FILE,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="complaints.xlsx"
    )


@app.get("/engineer-download")
def engineer_download(request: Request, token: str = ""):
    try:
        payload = verify_token(token, max_age=24 * 3600)
    except SignatureExpired:
        return HTMLResponse("<h2>Engineer link expired</h2>", status_code=403)
    except BadSignature:
        return HTMLResponse("<h2>Invalid engineer token</h2>", status_code=403)

    if payload.get("role") != "engineer_action":
        return HTMLResponse("<h2>Unauthorized</h2>", status_code=403)

    if not os.path.exists(COMPLAINTS_FILE):
        init_complaints_workbook()

    return FileResponse(
        COMPLAINTS_FILE,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="complaints.xlsx"
    )


@app.post("/resolve", response_class=HTMLResponse)
def resolve_complaint(
    request: Request,
    token: str = Form(...),
    status: str = Form("Resolved"),
    solution: str = Form("")
):
    try:
        payload = verify_token(token, max_age=24 * 3600)
    except SignatureExpired:
        return HTMLResponse("<h2>Resolution link expired</h2>", status_code=403)
    except BadSignature:
        return HTMLResponse("<h2>Invalid resolution token</h2>", status_code=403)

    if payload.get("role") != "engineer_action":
        return HTMLResponse("<h2>Unauthorized</h2>", status_code=403)

    complaint_id = payload.get("complaint_id", "")
    engineer_code = payload.get("engineer_code", "")

    complaint = get_complaint_by_id(complaint_id)

    if not complaint:
        return HTMLResponse("<h2>Complaint not found</h2>", status_code=404)

    try:
        if complaint.get("extra_issue_1") and not complaint.get("extra_solution_1"):
            ok, msg = save_extra_solution(complaint_id, solution.strip(), status)
            if not ok:
                return HTMLResponse(f"<h2>{msg}</h2>", status_code=400)
        elif complaint.get("extra_issue_2") and not complaint.get("extra_solution_2"):
            ok, msg = save_extra_solution(complaint_id, solution.strip(), status)
            if not ok:
                return HTMLResponse(f"<h2>{msg}</h2>", status_code=400)
        else:
            resolve_in_excel(complaint_id, status, solution.strip())
    except Exception as e:
        return HTMLResponse(f"<h2>{str(e)}</h2>", status_code=500)

    updated_complaint = get_complaint_by_id(complaint_id)

    customer_token = generate_token({
        "complaint_id": complaint_id,
        "role": "customer_ack"
    })
    acknowledgement_link = str(request.url_for("acknowledge_page", complaint_id=complaint_id)) + f"?token={customer_token}"

    if updated_complaint and updated_complaint.get("email"):
        mail_subject = f"Complaint {complaint_id} has been resolved"
        mail_body = f"""
Hello {updated_complaint.get('name', '')},

Your complaint has been resolved.

Complaint ID: {complaint_id}
Status: {status}
Resolution: {solution.strip()}

Please confirm whether the issue is fully resolved by clicking the link below:
{acknowledgement_link}

Regards,
Support Team
""".strip()

        send_email(updated_complaint.get("email"), mail_subject, mail_body)

    encoded_solution = quote_plus(solution.strip())
    return RedirectResponse(
        url=f"/resolved/{complaint_id}?status={status}&solution={encoded_solution}&token={token}",
        status_code=303
    )

@app.get("/acknowledge/{complaint_id}", response_class=HTMLResponse, name="acknowledge_page")
def acknowledge_page(request: Request, complaint_id: str, token: str = ""):
    complaint = get_complaint_by_id(complaint_id)
    if not complaint:
        return HTMLResponse("<h2>Complaint not found</h2>", status_code=404)

    try:
        payload = verify_token(token, max_age=3 * 24 * 3600)
    except (SignatureExpired, BadSignature):
        return HTMLResponse("<h2>Acknowledgement link expired or invalid</h2>", status_code=403)

    if payload.get("role") != "customer_ack" or payload.get("complaint_id") != complaint_id:
        return HTMLResponse("<h2>Unauthorized acknowledgement token</h2>", status_code=403)

    return templates.TemplateResponse(
        request=request,
        name="acknowledge.html",
        context={
            "complaint": complaint,
            "token": token,
            "error": ""
        }
    )


@app.post("/acknowledge/{complaint_id}/yes", response_class=HTMLResponse)
def acknowledge_yes(request: Request, complaint_id: str, token: str = Form(...)):
    complaint = get_complaint_by_id(complaint_id)
    if not complaint:
        return HTMLResponse("<h2>Complaint not found</h2>", status_code=404)

    try:
        payload = verify_token(token, max_age=3 * 24 * 3600)
    except (SignatureExpired, BadSignature):
        return HTMLResponse("<h2>Acknowledgement link expired or invalid</h2>", status_code=403)

    if payload.get("role") != "customer_ack" or payload.get("complaint_id") != complaint_id:
        return HTMLResponse("<h2>Unauthorized acknowledgement token</h2>", status_code=403)

    update_customer_acknowledgement(complaint_id, "Satisfied")

    updated = get_complaint_by_id(complaint_id)
    if updated and updated.get("email"):
        send_email(
            updated.get("email"),
            f"Complaint {complaint_id} closed",
            f"""
Hello {updated.get('name', '')},

Thank you for your confirmation.

Complaint ID: {complaint_id}
Status: Closed

Regards,
Support Team
""".strip()
        )

    return templates.TemplateResponse(
        request=request,
        name="ack_result.html",
        context={
            "title": "Thank you",
            "message": f"Your complaint {complaint_id} has been closed."
        }
    )


@app.post("/acknowledge/{complaint_id}/no", response_class=HTMLResponse)
def acknowledge_no(
    request: Request,
    complaint_id: str,
    token: str = Form(...),
    follow_up_issue: str = Form("")
):
    complaint = get_complaint_by_id(complaint_id)
    if not complaint:
        return HTMLResponse("<h2>Complaint not found</h2>", status_code=404)

    try:
        payload = verify_token(token, max_age=3 * 24 * 3600)
    except (SignatureExpired, BadSignature):
        return HTMLResponse("<h2>Acknowledgement link expired or invalid</h2>", status_code=403)

    if payload.get("role") != "customer_ack" or payload.get("complaint_id") != complaint_id:
        return HTMLResponse("<h2>Unauthorized acknowledgement token</h2>", status_code=403)

    if not follow_up_issue.strip():
        return templates.TemplateResponse(
            request=request,
            name="acknowledge.html",
            context={
                "complaint": complaint,
                "token": token,
                "error": "Please describe what issue still remains."
            }
        )
        

    extra1 = complaint.get("extra_issue_1")
    extra1_solution = complaint.get("extra_solution_1")
    extra2 = complaint.get("extra_issue_2")
    extra2_solution = complaint.get("extra_solution_2")

    if extra1 and extra1_solution and extra2 and extra2_solution:
        update_customer_acknowledgement(complaint_id, "Satisfied")
        updated = get_complaint_by_id(complaint_id)
        if updated and updated.get("email"):
            send_email(
                updated.get("email"),
                f"Complaint {complaint_id} closed",
                f"""
Hello {updated.get('name', '')},

Thank you for your feedback. This complaint has now been closed because the follow-up limit has been reached.

Complaint ID: {complaint_id}
Status: Closed

Please contact support directly for any further issues.

Regards,
Support Team
""".strip()
            )
        return templates.TemplateResponse(
            request=request,
            name="ack_result.html",
            context={
                "title": "Please contact support",
                "message": "We have reached the maximum follow-up limit for this complaint. The complaint is now closed and support will assist you directly."
            }
        )

    # Save the follow-up as the next extra issue in the workbook
    ok, msg = save_extra_issue(complaint_id, follow_up_issue.strip())
    if not ok:
        return HTMLResponse(f"<h2>{msg}</h2>", status_code=500)

    update_customer_acknowledgement(complaint_id, "Not Satisfied", follow_up_issue.strip())

    updated = get_complaint_by_id(complaint_id)
    engineer_code = updated.get("assigned_engineer", "")
    engineer_email = ENGINEER_EMAIL_MAP.get(engineer_code, "").strip()

    if engineer_email:
        engineer_token = generate_token({
            "complaint_id": complaint_id,
            "role": "engineer_action",
            "engineer_code": engineer_code
        })
        # Link engineer directly to the reopen page with token
        engineer_link = str(request.url_for("engineer_reopen", complaint_id=complaint_id)) + f"?token={engineer_token}"

        mail_subject = f"[{engineer_code}] Complaint reopened - {complaint_id}"
        mail_body = f"""
Complaint {complaint_id} has been reopened by the user.

Original issue: {updated.get('issue', '')}
Previous resolution: {updated.get('resolution', '')}
Customer follow-up issue: {follow_up_issue.strip()}

Open complaint:
{engineer_link}
""".strip()

        send_email(engineer_email, mail_subject, mail_body)

    return templates.TemplateResponse(
        request=request,
        name="ack_result.html",
        context={
            "title": "Feedback submitted",
            "message": "Your complaint has been reopened and sent back to the assigned engineer."
        }
    )

@app.get("/details-access/{complaint_id}", response_class=HTMLResponse)
def complaint_details_access_page(request: Request, complaint_id: str, error: str = ""):
    complaint = get_complaint_by_id(complaint_id)
    if not complaint:
        return HTMLResponse(content=f"<h2>Complaint ID {complaint_id} not found</h2>", status_code=404)

    return templates.TemplateResponse(
        request=request,
        name="admin_details_access.html",
        context={
            "complaint_id": complaint_id,
            "complaint": complaint,
            "error": error
        }
    )


@app.post("/details-access/{complaint_id}", response_class=HTMLResponse)
def complaint_details_access_submit(request: Request, complaint_id: str, password: str = Form("")):
    if is_valid_admin_password(password):
        return RedirectResponse(url=f"/complaint/{complaint_id}?password={password}", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="admin_details_access.html",
        context={
            "complaint_id": complaint_id,
            "complaint": get_complaint_by_id(complaint_id),
            "error": "Incorrect password. Only admin can view complaint details."
        }
    )


@app.get("/complaint/{complaint_id}", response_class=HTMLResponse)
def complaint_details(request: Request, complaint_id: str, password: str = "", token: str = ""):
    complaint = get_complaint_by_id(complaint_id)

    if not complaint:
        return HTMLResponse(
            content=f"<h2>Complaint ID {complaint_id} not found</h2>",
            status_code=404
        )

    if token:
        try:
            payload = verify_token(token, max_age=24 * 3600)
        except (SignatureExpired, BadSignature):
            return RedirectResponse(url=f"/details-access/{complaint_id}?error=unauthorized", status_code=303)

        if payload.get("role") == "engineer_action":
            return templates.TemplateResponse(
                request=request,
                name="response.html",
                context={
                    "complaint": complaint,
                    "token": token,
                    "engineer_code": payload.get("engineer_code", complaint.get("assigned_engineer", ""))
                }
            )

    if is_valid_admin_password(password):
        return templates.TemplateResponse(
            request=request,
            name="response.html",
            context={
                "complaint": complaint,
                "token": "",
                "engineer_code": complaint.get("assigned_engineer", "")
            }
        )

    return RedirectResponse(url=f"/details-access/{complaint_id}?error=unauthorized", status_code=303)